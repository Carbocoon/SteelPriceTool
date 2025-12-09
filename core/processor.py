import pandas as pd
from openpyxl import load_workbook
import re
from io import BytesIO
from core.rules import PricingRuleEngine

def normalize_data_columns(df):
    """
    标准化Data文件的列名，处理同义词
    """
    column_mapping = {
        "型号": ["规格", "型号", "规格型号", "规格/型号"],
        "支重_合并": ["支重", "重量", "公斤数", "支/公斤", "单重", "理重", "支/kg", "单支重（支数）", "单支重(支数)", "平均单支重量"],
        "单价": ["单价", "价格", "过磅价", "过磅", "金额", "单价/吨", "含税价"],
        "长度": ["长度", "米数", "长", "定尺"],
        "品名": ["品名", "货物名称", "商品名称", "名称", "货物", "存货"],
        "支数": ["支数", "数量", "总支数"], 
        "件数": ["件数", "件", "可用件数"],
        "支/件": ["单包支数", "支/件", "每件支数", "每包支数", "可用支数"]
    }
    
    # 反转映射：同义词 -> 标准名
    reverse_map = {}
    for std_col, synonyms in column_mapping.items():
        for syn in synonyms:
            reverse_map[syn] = std_col
            
    # 重命名列
    new_columns = {}
    extracted_product_name = None

    for col in df.columns:
        # 去除所有空白字符 (解决 "单包 支数" 这种带空格的表头)
        clean_col = re.sub(r'\s+', '', str(col))
        
        # 1. 精确匹配
        if clean_col in reverse_map:
            new_columns[col] = reverse_map[clean_col]
        else:
            # 2. 模糊匹配 (针对价格列)
            # 如果列名包含 "价格" 或 "金额" 或 "元/吨"，且不包含 "总价" (避免误判)，则映射为单价
            if ("价格" in clean_col or "金额" in clean_col or "元/吨" in clean_col or "单价" in clean_col or "含税价" in clean_col) and "总" not in clean_col:
                new_columns[col] = "单价"
            # 3. 特殊匹配：规格+品名 (如 "规格方管")
            # 如果列名以 "规格" 开头，且后面还有字，则认为是 "规格" + "品名"
            elif clean_col.startswith("规格") and len(clean_col) > 2:
                new_columns[col] = "型号"
                suffix = clean_col[2:].strip()
                # 只有当后缀包含汉字时，才认为是品名 (避免 "规格(mm)" 或 "规格/型号" 这种情况被误判)
                if re.search(r'[\u4e00-\u9fa5]', suffix):
                    extracted_product_name = suffix
            
    df = df.rename(columns=new_columns)
    
    # 如果从表头提取到了品名，自动填充到 "品名" 列
    if extracted_product_name:
        # 如果没有品名列，直接创建
        if "品名" not in df.columns:
            df["品名"] = extracted_product_name
        else:
            # 如果有品名列，填充其中的空值 (NaN 或 空字符串)
            df["品名"] = df["品名"].fillna(extracted_product_name)
            df["品名"] = df["品名"].replace(r'^\s*$', extracted_product_name, regex=True)
    
    # 处理重复列名（合并数据）
    if not df.columns.is_unique:
        # 获取所有唯一列名
        unique_cols = df.columns.unique()
        new_df = pd.DataFrame(index=df.index)
        
        for col in unique_cols:
            # 如果该列名对应多列（即有重复）
            if isinstance(df[col], pd.DataFrame):
                # 使用 bfill 合并（优先取非空值）
                # axis=1 对行操作，即横向合并这几列
                combined_col = df[col].bfill(axis=1).iloc[:, 0]
                new_df[col] = combined_col
            else:
                new_df[col] = df[col]
        df = new_df
        
    return df

def extract_embedded_rules(df):
    """
    从DataFrame中提取嵌入的加价规则
    寻找包含 "加价" 和 "标准" 的列，提取其下方的规则文本
    """
    found_rules = []
    # 遍历列名寻找关键词
    rule_cols = [c for c in df.columns if "加价" in str(c) and "标准" in str(c)]
    
    for col in rule_cols:
        # 找到列索引
        col_idx = df.columns.get_loc(col)
        
        # 遍历该列的非空行
        for idx, row in df.iterrows():
            desc = str(row[col]).strip()
            if not desc or desc.lower() == "nan":
                continue
                
            rule_str = desc
            
            # 尝试获取右边一列的值 (如果存在，且看起来像价格调整值)
            # 假设规则格式：Col A: "所有规格 0.7厚度", Col B: "+260元"
            if col_idx + 1 < len(df.columns):
                next_val = row.iloc[col_idx + 1]
                if pd.notna(next_val) and str(next_val).strip() != "":
                    val_str = str(next_val).strip()
                    # 简单的启发式：如果包含数字或+号，拼接到规则字符串
                    if any(c.isdigit() for c in val_str) or "+" in val_str:
                        rule_str += f" {val_str}"
            
            found_rules.append(rule_str)
            
    if found_rules:
        print(f"  发现嵌入的加价规则 ({len(found_rules)}条)")
        return "\n".join(found_rules)
    return ""

def parse_composite_weight(df):
    """
    处理复合支重列，如 "4.55 (144支)"
    分离出 支重 和 支数
    如果存在，将原 "支数" 列重命名为 "件数"，并计算 "支/件"
    """
    if "支重_合并" not in df.columns:
        return df
        
    # 正则匹配：数字 + 空格(可选) + 括号 + 数字 + 任意字符(如'支') + 反括号
    pattern = r"^(\d+\.?\d*)\s*[（(](\d+).*?[)）]$"
    
    def extract_info(val):
        if pd.isna(val):
            return None, None
        s = str(val).strip()
        match = re.match(pattern, s)
        if match:
            return float(match.group(1)), int(match.group(2))
        return None, None

    extracted = df["支重_合并"].apply(extract_info)
    
    # 检查是否提取到了数据 (至少有一行匹配)
    valid_count = extracted.apply(lambda x: x[0] is not None).sum()
    
    if valid_count == 0:
        return df
        
    print(f"  检测到复合支重格式 (如 '4.55 (144支)')，共 {valid_count} 行，正在拆分...")
    
    # 1. 如果存在 "支数" 列，说明它实际上是 "件数" (根据用户描述)
    # 先重命名，腾出 "支数" 这个名字给真正的支数
    if "支数" in df.columns:
        df = df.rename(columns={"支数": "件数"})
        print("  已将原 '支数' 列重命名为 '件数'")
        
    # 2. 提取数据并更新列
    weights = extracted.apply(lambda x: x[0])
    counts = extracted.apply(lambda x: x[1])
    
    # 更新 支重_合并 (只更新匹配的行)
    mask = weights.notna()
    df.loc[mask, "支重_合并"] = weights[mask]
    
    # 更新/创建 支数
    if "支数" not in df.columns:
        df["支数"] = pd.NA
    df.loc[mask, "支数"] = counts[mask]
    
    # 3. 计算 支/件 (如果 件数 存在)
    if "件数" in df.columns:
        # 确保是数值类型
        df["件数"] = pd.to_numeric(df["件数"], errors='coerce')
        df["支数"] = pd.to_numeric(df["支数"], errors='coerce')
        
        # 计算
        df["支/件"] = df["支数"] / df["件数"]
        # 处理除以0或空值的情况
        df["支/件"] = df["支/件"].fillna(0).replace(float('inf'), 0)
        print("  已计算 '支/件' = 支数 / 件数")
        
    return df

def extract_product_info(product_name):
    """
    从混合字符串中分离品名和型号
    支持两种格式：
    1. 品名在前，型号在后 (e.g. "槽钢16#") -> Name="槽钢", Model="16#"
    2. 型号在前，品名在后 (e.g. "16#槽钢") -> Name="槽钢", Model="16#"
    """
    if pd.isna(product_name) or str(product_name).strip() == "":
        return "", ""
    
    # 使用正则去除所有空白字符
    s = re.sub(r'\s+', '', str(product_name))
    
    # 1. 寻找第一个汉字的位置
    match = re.search(r'[\u4e00-\u9fa5]', s)
    
    if not match:
        # 没有汉字，假设全是型号（或者无法识别）
        return "", s
        
    first_cn_idx = match.start()
    
    if first_cn_idx == 0:
        # Case A: 以汉字开头 -> 品名在前，型号在后
        # 提取开头的连续汉字作为品名
        # e.g. "槽钢16#" -> Name="槽钢", Model="16#"
        cn_match = re.match(r'^([\u4e00-\u9fa5]+)', s)
        if cn_match:
            name_part = cn_match.group(1)
            model_part = s[len(name_part):]
            return name_part, model_part.replace("#", "").strip()
        else:
            return s, ""
            
    else:
        # Case B: 以非汉字开头 -> 型号在前，品名在后
        # e.g. "16#槽钢" -> Model="16#", Name="槽钢"
        model_part = s[:first_cn_idx]
        name_part = s[first_cn_idx:]
        
        return name_part, model_part.replace("#", "").strip()

def preprocess_prices(df, default_base=3300, rule_engine=None, price_mode="auto", base_price_strategy="auto_identify"):
    """
    预处理价格：按【品名】分组，寻找组内基价（完整大数），计算加价。
    逻辑：
    1. 识别每行的价格类型（完整价格 vs 加价）。
    2. 按品名分组，找到该品名下的基价（通常是唯一的完整大数）。
    3. 计算最终价格：基价 + 加价。
    """
    if "单价" not in df.columns:
        return df
    
    # 临时辅助函数：解析价格字符串
    def parse_price_str(val):
        if pd.isna(val) or str(val).strip() == "":
            return None, False # val, is_add
        
        # 使用正则去除所有空白字符，解决 "2 0" 或 "2\t0" 被识别为 2 的问题
        s = re.sub(r'\s+', '', str(val))
        # 去除 "元"、"元/吨" 等后缀，防止影响数字提取或被误判
        s = s.replace("元/吨", "").replace("元", "")
        
        # 提取数字
        num_match = re.search(r'(\d+\.?\d*)', s)
        if not num_match:
            return None, False
        
        num = float(num_match.group(1))
        
        # 判定是否为加价：
        # 1. 显式包含 "加" 或 "+"
        # 2. 数值较小（< 1000，通常钢材单价在3000+）
        is_explicit_add = "加" in s or "+" in s
        is_small_val = num < 1000
        
        is_add = is_explicit_add or is_small_val
        return num, is_add

    # 1. 解析所有价格
    parsed = df["单价"].apply(parse_price_str)
    df["_temp_val"] = [x[0] for x in parsed]
    df["_is_add"] = [x[1] for x in parsed]

    # --- 模式分支 ---
    if price_mode == "direct":
        # 直接读取模式：直接使用解析出的数值，忽略加价标记
        # 仍然需要处理合并单元格的填充问题（如果品名存在）
        if "品名" in df.columns:
             df["_final_price"] = df.groupby("品名")["_temp_val"].ffill()
        else:
             df["_final_price"] = df["_temp_val"]
        return df

    if "品名" not in df.columns:
        # 如果没有品名列，创建一个临时品名列用于分组，确保计算能进行
        print("  提示：未找到“品名”列，将把所有数据视为同一品名进行处理")
        df["品名"] = "默认品名"

    # 1. 处理合并单元格：按品名分组，对单价进行向前填充（ffill）
    # 解决“合并单元格要填一样”的问题，同时避免跨品名污染
    df["单价"] = df.groupby("品名")["单价"].ffill()
    
    # 2. 按品名分组计算
    # 创建一个新的列存储最终价格
    final_prices = []
    
    # 按品名分组处理
    for name, group in df.groupby("品名"):
        # 寻找该组内的基价：取非加价的最大值（假设基价是完整价格）
        # 如果组内有多个完整价格，通常取出现频率最高或最大的，这里取最大值作为基准
        
        group_base = None
        
        # --- 规则引擎：基价查找 ---
        if rule_engine and price_mode == "rule":
            # 尝试从规则中获取基价型号
            base_model_name = rule_engine.get_base_model(str(name))
            if base_model_name:
                # 在当前组中寻找该型号的价格
                # 假设型号在 "型号_临时" 列中
                if "型号_临时" in group.columns:
                    # 模糊匹配型号
                    base_row = group[group["型号_临时"].astype(str).str.contains(base_model_name, na=False)]
                    # 且必须是完整价格（非加价）
                    base_row = base_row[~base_row["_is_add"]]
                    
                    if not base_row.empty:
                        group_base = base_row["_temp_val"].max()
                        # print(f"  [{name}] 使用规则基价型号 {base_model_name}: {group_base}")

        if group_base is None:
            # Auto 模式下的基价策略 (合并了原 Embedded 模式逻辑)
            # 如果策略是 fixed，直接用 default_base
            # 如果策略是 auto_identify (默认)，尝试在组内找
            
            if base_price_strategy == "fixed":
                group_base = default_base
            else:
                # auto_identify: 尝试在组内找
                base_candidates = group[~group["_is_add"]]["_temp_val"]
                if not base_candidates.empty:
                    group_base = base_candidates.max()
                else:
                    group_base = default_base
            
        # 计算该组所有行的最终价格
        for idx, row in group.iterrows():
            val = row["_temp_val"]
            is_add = row["_is_add"]
            
            if pd.isna(val):
                # 如果ffill后仍为空，说明该品名下确实无价格，保持为空
                final_prices.append((idx, None))
                continue
            
            # --- 规则引擎：额外加价计算 ---
            rule_adjustment = 0
            # 在 rule 模式下，或者在 auto 模式下且有规则引擎时，都进行计算
            if rule_engine and (price_mode == "rule" or price_mode == "auto"):
                
                # 内部辅助函数：安全转换浮点数（处理 "1. 59" 这种带空格的情况）
                def safe_float_val(v):
                    try:
                        if pd.isna(v): return 0
                        # 去除所有空白字符
                        s = re.sub(r'\s+', '', str(v))
                        return float(s)
                    except:
                        return 0

                # 准备行数据供规则引擎使用
                row_data = {
                    "product": str(row["品名"]),
                    "model": str(row.get("型号_临时", "")),
                    "length": safe_float_val(row.get("长度")),
                    "weight": safe_float_val(row.get("支重_合并")),
                    "diff": safe_float_val(row.get("负差")),
                    "raw_spec": str(row.get("品名", "")) + str(row.get("型号_临时", "")) # 简单拼接用于全规格匹配
                }
                
                rule_adjustment = rule_engine.calculate_adjustment(row_data)

            if is_add:
                final_price = group_base + val + rule_adjustment
            else:
                final_price = val + rule_adjustment # 已经是完整价格，但也可能适用规则加价？通常规则是针对基价的加价
                # 如果是完整价格，通常意味着它已经是最终价，但如果规则说 "40角加价120"，这通常是在基价基础上的加价
                # 如果当前行是完整价格，是否还要加？
                # 假设：如果当前行是完整价格，它可能就是基价本身，或者是特殊定价。
                # 如果规则是 "统一加价"，那应该加。
                # 这里的逻辑比较模糊，暂定：如果是完整价格，也加上规则调整值（假设规则是额外的）
            
            final_prices.append((idx, final_price))
    
    # 3. 将计算结果填回 DataFrame
    # 因为groupby打乱了顺序，需要按索引映射回去
    price_map = dict(final_prices)
    df["_final_price"] = df.index.map(price_map)
    
    return df

def generate_preview_dataframe(data_file, model_file, default_base_price=3300, rule_text=None, price_mode="auto", global_product_name=None, base_price_strategy="auto_identify", preloaded_data=None, user_mapping_df=None):
    """
    生成预览用的DataFrame (包含完整的数据处理和映射逻辑)
    """
    # ---------------------- 1. 读取Data文件 ----------------------
    print("【预览生成】读取Data文件...")
    if preloaded_data is not None:
        combined_data = preloaded_data
    else:
        if hasattr(data_file, 'seek'): data_file.seek(0)
        try:
            all_sheets = pd.read_excel(data_file, sheet_name=None, header=0, keep_default_na=False)
        except Exception as e:
            raise ValueError(f"读取Data文件失败: {str(e)}")

        data_sheets = []
        for sheet_name, sheet_data in all_sheets.items():
            sheet_data = sheet_data.dropna(how='all')
            sheet_data = normalize_data_columns(sheet_data)
            sheet_data = parse_composite_weight(sheet_data)
            if "品名" in sheet_data.columns:
                sheet_data["品名"] = sheet_data["品名"].replace("", pd.NA).ffill()
            data_sheets.append(sheet_data)
        
        if not data_sheets:
            raise ValueError("未读取到任何Data数据")
        combined_data = pd.concat(data_sheets, ignore_index=True)

    # ---------------------- 0. 初始化规则引擎 ----------------------
    rule_engine = None
    if rule_text:
        rule_engine = PricingRuleEngine()
        rule_engine.parse_rules(rule_text)

    # ---------------------- 2. 获取Model Headers ----------------------
    print("【预览生成】获取Model表头...")
    if hasattr(model_file, 'seek'): model_file.seek(0)
    wb = load_workbook(model_file)
    ws = wb.active
    
    # 模拟 process_files 的列处理逻辑 (删除第一列)
    start_col = 2 if ws.max_column >= 1 else 1
    
    model_headers = []
    for col in range(start_col, ws.max_column + 1):
        val1 = ws.cell(row=1, column=col).value
        val2 = ws.cell(row=2, column=col).value
        header_val = val2 if (val2 and str(val2).strip()) else val1
        if header_val and str(header_val).strip():
            model_headers.append(str(header_val).strip())
        else:
            model_headers.append(f"空列_{col}")
            
    for i, h in enumerate(model_headers):
        if "规格" in h or "型号" in h:
            model_headers[i] = "型号"
            break
    
    for i, h in enumerate(model_headers):
        if "价格" in h or "过磅" in h or "单价" in h or "含税价" in h:
            model_headers[i] = "过磅"
            break
            
    # 确保列名唯一 (Streamlit data_editor 要求)
    seen = {}
    for i, h in enumerate(model_headers):
        if h not in seen:
            seen[h] = 1
        else:
            count = seen[h]
            seen[h] += 1
            model_headers[i] = f"{h}_{count}"

    # ---------------------- 4. Data预处理 ----------------------
    print("【预览生成】Data预处理...")
    key_cols_to_check = [c for c in ["型号", "单价", "支重_合并", "长度", "支数"] if c in combined_data.columns]
    if key_cols_to_check:
        for col in key_cols_to_check:
            combined_data[col] = combined_data[col].replace(r'^\s*$', pd.NA, regex=True)
        combined_data = combined_data.dropna(subset=key_cols_to_check, how='all')

    if global_product_name:
        if "品名" not in combined_data.columns:
            combined_data["品名"] = global_product_name
        else:
            combined_data["品名"] = combined_data["品名"].replace("", pd.NA).fillna(global_product_name)

    if "型号" in combined_data.columns:
        combined_data["型号_临时"] = combined_data["型号"].astype(str).apply(lambda x: re.sub(r'\s+', '', x))
        if "品名" in combined_data.columns:
             combined_data["品名_清洗"] = combined_data["品名"].astype(str).apply(lambda x: re.sub(r'\s+', '', x))
        else:
             combined_data["品名_清洗"] = global_product_name if global_product_name else ""
    elif "品名" in combined_data.columns:
        extracted_info = combined_data["品名"].apply(extract_product_info)
        combined_data["型号_临时"] = [x[1] for x in extracted_info]
        combined_data["品名_清洗"] = [x[0] for x in extracted_info]
    else:
        if price_mode == "direct":
             combined_data["型号_临时"] = ""
             combined_data["品名_清洗"] = ""
        else:
             raise ValueError("Data文件缺少“品名”或“规格/型号”列")

    if "支重_合并" not in combined_data.columns:
        combined_data["支重_合并"] = pd.NA

    # ---------------------- 4.5 合并用户映射 & 价格计算 ----------------------
    if user_mapping_df is not None and not user_mapping_df.empty:
        try:
            mapping_renamed = user_mapping_df.rename(columns={"品名": "品名_清洗", "型号": "型号_临时"})
            for df_obj in [combined_data, mapping_renamed]:
                for col in ["品名_清洗", "型号_临时"]:
                    if col in df_obj.columns:
                        df_obj[col] = df_obj[col].fillna("").astype(str).replace("nan", "")
            
            cols_to_merge = [c for c in mapping_renamed.columns if c not in ["品名_清洗", "型号_临时"]]
            cols_to_drop = [c for c in cols_to_merge if c in combined_data.columns]
            if cols_to_drop:
                combined_data = combined_data.drop(columns=cols_to_drop)

            combined_data = pd.merge(combined_data, mapping_renamed, on=["品名_清洗", "型号_临时"], how="left")
        except Exception as e:
            print(f"  合并用户自定义属性失败: {str(e)}")

    combined_data = preprocess_prices(combined_data, default_base=default_base_price, rule_engine=rule_engine, price_mode=price_mode, base_price_strategy=base_price_strategy)

    # 排序
    has_price = "_final_price" in combined_data.columns
    def is_valid_row(row):
        model_valid = pd.notna(row.get("型号_临时")) and str(row.get("型号_临时")).strip() != ""
        price_valid = False
        if has_price:
            price_valid = pd.notna(row["_final_price"])
        return model_valid or price_valid

    combined_data = combined_data[combined_data.apply(is_valid_row, axis=1)]

    combined_data["_name_sort_val"] = combined_data["品名_清洗"].apply(lambda x: str(x).encode('gbk', errors='ignore'))
    combined_data["_model_sort_val"] = combined_data["型号_临时"].apply(lambda x: float(re.search(r"(\d+(\.\d+)?)", str(x)).group(1)) if re.search(r"(\d+(\.\d+)?)", str(x)) else 0)
    combined_data["_weight_sort_val"] = pd.to_numeric(combined_data["支重_合并"], errors='coerce').fillna(0)
    
    if "支/件" not in combined_data.columns:
        combined_data["支/件"] = pd.NA
    if "支数" in combined_data.columns and "件数" in combined_data.columns:
         mask = combined_data["支/件"].isna()
         c_counts = pd.to_numeric(combined_data["件数"], errors='coerce')
         c_pieces = pd.to_numeric(combined_data["支数"], errors='coerce')
         calculated = c_pieces / c_counts
         combined_data.loc[mask, "支/件"] = calculated[mask].fillna(0).replace(float('inf'), 0)

    combined_data = combined_data.sort_values(by=["_name_sort_val", "_model_sort_val", "_weight_sort_val"], ascending=[True, True, True]).reset_index(drop=True)

    # ---------------------- 5. 构建 DataFrame ----------------------
    field_mapping = {
        "品名_清洗": ("品名", lambda x: x if x else global_product_name),
        "厂家": ("品牌/厂家", lambda x: x),
        "长度": ("长度/mm", lambda x: x * 1000 if str(x).replace(".", "").isdigit() else x),
        "支/件": ("支/件", lambda x: x if str(x).replace(".", "").isdigit() else x),
        "支重_合并": ("支重/kg", lambda x: x),
        "负差": ("负差", lambda x: x),
        "_final_price": ("过磅", lambda x: x),
        "型号_临时": ("型号", lambda x: x)
    }

    final_rows = []
    for _, data_row in combined_data.iterrows():
        row_data = {}
        # 1. 映射字段
        for data_col, (target_col, process_func) in field_mapping.items():
            if data_col in combined_data.columns:
                val = data_row[data_col]
                if isinstance(val, pd.Series): val = val.dropna().iloc[0] if not val.dropna().empty else None
                
                is_valid_val = (pd.notna(val) and str(val).strip() != "" and str(val).strip().lower() != "nan")
                if target_col == "品名": is_valid_val = True

                if is_valid_val and target_col in model_headers:
                    row_data[target_col] = process_func(val)
        
        # 2. 剩余字段
        for target_col in model_headers:
            if target_col not in row_data:
                if target_col in combined_data.columns:
                    val = data_row[target_col]
                    if pd.notna(val):
                        row_data[target_col] = val
                else:
                    row_data[target_col] = None # 显式置空
        
        # 按 model_headers 顺序排列
        ordered_row = [row_data.get(col) for col in model_headers]
        final_rows.append(ordered_row)

    final_df = pd.DataFrame(final_rows, columns=model_headers)
    return final_df

def analyze_data_and_model(data_file, model_file, global_product_name=None):
    # 兼容旧接口，但返回完整预览
    df = generate_preview_dataframe(data_file, model_file, global_product_name=global_product_name)
    return {"preview_df": df}

def process_files(data_file, model_file, default_base_price=3300, rule_text=None, price_mode="auto", global_product_name=None, base_price_strategy="auto_identify", preloaded_data=None, user_mapping_df=None, final_df=None):
    
    # 如果没有提供 final_df，则生成它
    if final_df is None:
        final_df = generate_preview_dataframe(data_file, model_file, default_base_price, rule_text, price_mode, global_product_name, base_price_strategy, preloaded_data, user_mapping_df)

    # ---------------------- 6. 写入 Excel ----------------------
    print("【生成文件】写入Excel...")
    if hasattr(model_file, 'seek'): model_file.seek(0)
    wb = load_workbook(model_file)
    ws = wb.active
    
    if ws.max_column >= 1:
        ws.delete_cols(1)
    
    ranges_to_unmerge = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row >= 3:
            ranges_to_unmerge.append(merged_range)
    for merged_range in ranges_to_unmerge:
        ws.unmerge_cells(str(merged_range))

    # 写入数据
    start_row = 3
    # final_df 的列顺序应该已经匹配了 (因为它是根据 model_headers 生成的)
    # 但为了保险，我们还是按列索引写
    
    # 将 DataFrame 转换为 list of lists
    data_values = final_df.values.tolist()
    
    for r_idx, row_val in enumerate(data_values):
        for c_idx, val in enumerate(row_val):
            # 只有非空值才写入？
            # 用户要求：没填的置空。
            # 如果 val 是 None 或 NaN，我们应该写入空字符串还是跳过？
            # 如果跳过，模板里的旧数据会保留。
            # 所以应该写入。如果 val 是 None，写入 ""。
            
            write_val = val
            if pd.isna(val):
                write_val = ""
            
            ws.cell(row=start_row + r_idx, column=c_idx + 1, value=write_val)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
